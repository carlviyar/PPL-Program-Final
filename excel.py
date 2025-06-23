from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime

def import_excel_data():
    workbook = load_workbook(filename="PPL Prices Reference CLEAN.xlsx", data_only=True)

    sheet = workbook.active

    # every "column" has 3 sub-columns: test code, desc, price
    # note: max_column is one-based
    numCols = sheet.max_column // 3

    # ignoring the first column, which contains base price and instructions
    numCols -= 1

    # labs: List of lists. Each list within the larger list contains the labs for that form column. i.e. labs[0] is a list of labs starting from 249 to 483.
    # lab: each lab is an object with the fields: test_code, description, price
    labs = []
    # lab_info: lab_price[test_code] = {"test_code", "description", "price"}
    lab_info = {}

    for colIndex in range(0, numCols):
        # colIndex: 0, 1, 2 [zero-based]
        # minCol:   3, 6, 9 [zero-based, inclusive]
        # maxCol:   5, 8, 11 [zero-based, inclusive]
        minCol = (colIndex+1)*3
        maxCol = minCol+2
        labs.insert(colIndex, [])
        for row in sheet.iter_rows(min_row=2,
                                    min_col=minCol+1,  # min_col is one-based index
                                    max_col=maxCol+1,  # max_col is one-based index
                                    values_only=True):
            if row[0] is None:
                continue
            lab = {
                "test_code": row[0],
                "description": row[1],
                "price": row[2]
            }
            lab_info[row[0]] = lab 
            labs[colIndex].append(lab)

    return labs, lab_info

# invoice_info: {name, notes, ordered_labs, custom_labs, prepaid}
def create_workbook(invoice_info):
    wb = load_workbook(filename="PPL Form Template CLEAN.xlsx", data_only=True)
    ws = wb.active

    # output_file name
    full_name = invoice_info["name"]
    output_file = output_file_name(full_name)

    # set name
    ws["B7"] = full_name

    # set date
    ws["B8"] = datetime.datetime.today()
    ws["B8"].number_format = "mm/dd/yyyy"

    # set notes
    ws["B9"] = invoice_info["notes"]

    # set base price
    base_price = get_base_price()
    ws["C13"] = "Base Price"
    ws["D13"] = base_price

    current_row = 14
    # set list of labs
    for lab in invoice_info["ordered_labs"]:
        # test code
        test_code_cell = ws.cell(row=current_row, column=2)
        test_code_cell.value = int(lab["test_code"])

        # description
        desc_cell = ws.cell(row=current_row, column=3)
        desc_cell.value = lab["description"]

        # price
        price = round(float(lab["price"]), 2)
        price_cell = ws.cell(row=current_row, column=4)
        price_cell.value = price
        
        current_row += 1
    
    current_row += 1
    # set total
    total_label_cell = ws.cell(row=current_row, column=3)
    total_label_cell.value = "TOTAL:"
    total_label_cell.font = Font(bold=True, name="Calisto MT", size=12)

    total_cost_cell = ws.cell(row=current_row, column=4)
    total_cost_cell.value = f"=SUM(D13:D{total_cost_cell.row-1})"

    current_row += 1
    # set prepaid
    prepaid_label_cell = ws.cell(row=current_row, column=3)
    prepaid_label_cell.value = "PREPAID:"
    prepaid_label_cell.font = Font(bold=True, name="Calisto MT", size=12)

    prepaid_value_cell = ws.cell(row=current_row, column=4)
    prepaid_value_cell.value = invoice_info["prepaid"]

    # finalized
    wb.save(output_file)

def output_file_name(full_name):
    names = full_name.split()
    today = datetime.datetime.today().strftime('%m-%d-%y')
    if len(names) == 0:
        return f"PPL READY {today}.xlsx"
    filepath = '.'
    if len(names) > 1:
        last_name = names[-1]
        first_name = ' '.join(names[0:-1])
        return f"{filepath}/{last_name}, {first_name} {today} PPL READY.xlsx"
    else:
        last_name = names[0]
        return f"{filepath}/{last_name} {today} PPL READY.xlsx"
    
def get_base_price():
    """
    Relies that the base price is on C2

    Returns price as float
    """
    workbook = load_workbook(filename="PPL Prices Reference CLEAN.xlsx", data_only=True)
    ws = workbook.active

    price = float(ws["C2"].value)
    return price
